#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys
import os

def check_excel():
    file_path = "CARRY X Doeat 주문현황.xlsx"

    if not os.path.exists(file_path):
        print(f"ERROR: File not found: {file_path}")
        return False

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        print("Excel file analysis:")
        print(f"Sheet name: {sheet.title}")
        print(f"Total rows: {sheet.max_row}")
        print(f"Total columns: {sheet.max_column}")

        print("\nColumn headers:")
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header)
            print(f"{col:2d}. {header}")

        print(f"\nSample data (first 3 rows):")
        for row in range(1, min(4, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(6, sheet.max_column + 1)):  # 처음 5개 컬럼만
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"Row {row}: {row_data}")

        workbook.close()
        return True, headers

    except Exception as e:
        print(f"ERROR: {str(e)}")
        return False, []

if __name__ == "__main__":
    success, headers = check_excel()
    if not success:
        print("Excel analysis failed - stopping work.")
        sys.exit(1)
    else:
        print("Excel analysis completed successfully.")