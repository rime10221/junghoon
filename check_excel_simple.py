#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys
import os

def check_excel_with_openpyxl():
    file_path = "CARRY X Doeat 주문현황.xlsx"

    if not os.path.exists(file_path):
        print(f"ERROR: 파일을 찾을 수 없습니다: {file_path}")
        return False

    try:
        # Excel 파일 읽기 (openpyxl 직접 사용)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        print("=" * 50)
        print("📊 EXCEL 파일 구조 분석 결과")
        print("=" * 50)
        print(f"✅ Excel 파일 읽기 성공!")
        print(f"📋 시트명: {sheet.title}")
        print(f"📋 전체 행 수: {sheet.max_row}")
        print(f"📋 전체 컬럼 수: {sheet.max_column}")

        # 헤더 (첫 번째 행) 읽기
        print("\n🔍 컬럼 목록 (헤더):")
        headers = []
        for col in range(1, sheet.max_column + 1):
            header = sheet.cell(row=1, column=col).value
            headers.append(header)
            print(f"  {col:2d}. {header}")

        # 첫 3행의 데이터 미리보기
        print(f"\n📄 첫 3행 데이터:")
        for row in range(1, min(4, sheet.max_row + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            print(f"  행 {row}: {row_data}")

        # 컬럼별 샘플 데이터 (2번째 행)
        print(f"\n🔍 컬럼별 샘플 데이터 (2번째 행):")
        if sheet.max_row >= 2:
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                sample = sheet.cell(row=2, column=col).value
                print(f"  {header}: {sample}")

        workbook.close()
        return True

    except Exception as e:
        print(f"ERROR: Excel 파일 읽기 실패: {str(e)}")
        return False

if __name__ == "__main__":
    success = check_excel_with_openpyxl()
    if not success:
        print("\n❌ Excel 파일 분석 실패 - 작업을 중단합니다.")
        sys.exit(1)
    else:
        print("\n✅ Excel 파일 분석 완료")