#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import sys
import os

def check_excel_values():
    file_path = "CARRY X Doeat 주문현황.xlsx"

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)  # data_only=True로 값만 읽기
        sheet = workbook.active

        print("Excel file real values analysis:")
        print(f"Sheet name: {sheet.title}")
        print(f"Total rows: {sheet.max_row}")
        print(f"Total columns: {sheet.max_column}")

        # 실제 헤더 추출 시도
        print("\nTrying to extract real column names...")

        # 여러 행을 체크해서 실제 데이터 찾기
        real_headers = []
        data_start_row = None

        for row in range(1, min(10, sheet.max_row + 1)):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)

            print(f"Row {row}: {row_data[:5]}...")  # 처음 5개만 출력

            # 데이터가 있는 행 찾기
            if any(val is not None and str(val).strip() != "" for val in row_data):
                if data_start_row is None:
                    data_start_row = row
                    # 첫 번째 데이터 행을 헤더로 간주
                    real_headers = [str(val) if val is not None else f"Column_{i}" for i, val in enumerate(row_data, 1)]

        print(f"\nDetected data start row: {data_start_row}")
        print("Extracted headers:")
        for i, header in enumerate(real_headers, 1):
            print(f"{i:2d}. {header}")

        # 실제 데이터 샘플 확인
        if data_start_row and data_start_row < sheet.max_row:
            print(f"\nSample data (3 rows starting from row {data_start_row + 1}):")
            for row in range(data_start_row + 1, min(data_start_row + 4, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else "")
                print(f"Row {row}: {row_data}")

        workbook.close()
        return True, real_headers, data_start_row

    except Exception as e:
        print(f"ERROR: {str(e)}")
        return False, [], None

if __name__ == "__main__":
    success, headers, start_row = check_excel_values()
    if not success:
        print("\nERROR: Excel analysis failed - stopping work.")
        sys.exit(1)
    else:
        print(f"\nSUCCESS: Excel analysis completed. Found {len(headers)} columns, data starts at row {start_row}.")